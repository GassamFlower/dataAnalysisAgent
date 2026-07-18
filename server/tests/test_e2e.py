"""端到端测试：完整流程 + 异常分支。

运行方式（需后端服务运行在 localhost:8000）：
    $env:PYTHONIOENCODING="utf-8"; python tests/test_e2e.py

测试组织：
- 主流程（happy path）：覆盖创建→体检→假设→生成→分析→获取报告→导出 Word/Excel→清理
- 异常分支（anomalies）：认证 / 状态守卫 / 404 / 参数校验，不依赖 LLM

注意：主流程会真实调用 DeepSeek LLM（R1~R4），耗时较长（约 30~60s）。
异常分支不创建假设、不调用 LLM，可独立快速运行。
"""
import io
import json
import sys
import time
import uuid
from typing import Optional

import requests

# Windows PowerShell 5 默认编码非 UTF-8，强制标准输出为 UTF-8 以避免中文乱码
if sys.platform == "win32":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8")

BASE_URL = "http://localhost:8000/api/v1"
HEALTH_URL = "http://localhost:8000/health"
HEADERS = {"Authorization": "Bearer dev-token", "Content-Type": "application/json"}

# 测试用题目（覆盖正向/反向表述，便于体检识别维度与反向题）
SAMPLE_QUESTIONS = """
1. 我对目前的工作环境感到满意
2. 我的工作内容让我感到有成就感
3. 我与同事之间的关系融洽
4. 我不喜欢目前的工作节奏
5. 我的薪资待遇符合我的期望
6. 我有清晰的职业发展路径
7. 我的工作压力在可接受范围内
8. 我不认同公司的管理方式
9. 我的工作与生活平衡良好
10. 我对公司的未来充满信心
"""

SAMPLE_HYPOTHESIS = "工作满意度正向影响职业发展，工作压力负向影响工作满意度"


# ============================================================
# 通用工具
# ============================================================

def _print_section(title: str) -> None:
    print("\n" + "=" * 60)
    print(title)
    print("=" * 60)


def _print_response(resp: requests.Response, show_body: bool = True) -> None:
    print(f"  状态码: {resp.status_code}")
    if show_body:
        try:
            print(f"  响应: {json.dumps(resp.json(), ensure_ascii=False, indent=2)[:500]}")
        except Exception:
            print(f"  响应（非 JSON）: {resp.text[:200]}")


def _create_draft_project(name_suffix: str = "") -> str:
    """创建一个 draft 项目，返回 project_id。供异常分支测试使用。"""
    resp = requests.post(
        f"{BASE_URL}/projects/",
        json={"name": f"异常测试-{name_suffix}-{uuid.uuid4().hex[:8]}"},
        headers=HEADERS,
    )
    assert resp.status_code == 201, f"创建项目失败: {resp.status_code} {resp.text}"
    return resp.json()["data"]["id"]


def _delete_project(project_id: str) -> None:
    """删除项目（清理）。"""
    requests.delete(f"{BASE_URL}/projects/{project_id}", headers=HEADERS)


def _activate_dev_subscription(plan_type: str = "single") -> None:
    """为 dev 用户激活指定套餐，供需要付费权限的异常分支测试使用。"""
    resp = requests.post(
        f"{BASE_URL}/payment/orders",
        json={"plan_type": plan_type},
        headers=HEADERS,
    )
    assert resp.status_code == 200, f"创建订单失败: {resp.status_code} {resp.text}"
    order_id = resp.json()["data"]["id"]

    notify_resp = requests.post(
        f"{BASE_URL}/payment/orders/{order_id}/notify",
        json={
            "channel": "wechat",
            "transaction_id": f"e2e-{uuid.uuid4().hex[:8]}",
            "status": "success",
        },
        headers=HEADERS,
    )
    assert notify_resp.status_code == 200, f"支付回调失败: {notify_resp.status_code} {notify_resp.text}"


# ============================================================
# 主流程（happy path）
# ============================================================

def test_health() -> None:
    """1. 健康检查。"""
    _print_section("1. 健康检查")
    resp = requests.get(HEALTH_URL)
    _print_response(resp, show_body=False)
    assert resp.status_code == 200
    data = resp.json()
    assert data["code"] == 0
    assert data["data"]["status"] == "ok"
    print("[OK] 健康检查通过")


def test_create_project() -> str:
    """2. 创建项目。"""
    _print_section("2. 创建项目")
    resp = requests.post(
        f"{BASE_URL}/projects/",
        json={"name": "E2E 测试 - 工作满意度调查"},
        headers=HEADERS,
    )
    _print_response(resp)
    assert resp.status_code == 201
    project_id = resp.json()["data"]["id"]
    print(f"[OK] 项目创建成功，ID: {project_id}")
    return project_id


def test_inspect_questions(project_id: str) -> None:
    """3. 题目体检（调用 LLM R1~R3）。"""
    _print_section("3. 题目体检")
    print("  调用 LLM，预计 5~15s...")
    t0 = time.time()
    resp = requests.post(
        f"{BASE_URL}/questionnaire/inspect?project_id={project_id}",
        json={"text": SAMPLE_QUESTIONS},
        headers=HEADERS,
        timeout=120,
    )
    print(f"  耗时: {time.time() - t0:.1f}s")
    _print_response(resp)
    assert resp.status_code == 200, f"题目体检失败: {resp.text}"
    data = resp.json()["data"]
    assert "questions" in data
    assert len(data["questions"]) == 10, f"题目数不对: {len(data['questions'])}"
    print(f"[OK] 题目体检完成，识别 {len(data['questions'])} 题")


def test_create_hypothesis(project_id: str) -> str:
    """4. 创建假设（调用 LLM 解析路径）。"""
    _print_section("4. 创建假设")
    print("  调用 LLM，预计 3~10s...")
    t0 = time.time()
    resp = requests.post(
        f"{BASE_URL}/simulation/{project_id}/hypothesis",
        json={"raw_text": SAMPLE_HYPOTHESIS},
        headers=HEADERS,
        timeout=120,
    )
    print(f"  耗时: {time.time() - t0:.1f}s")
    _print_response(resp)
    assert resp.status_code == 200, f"创建假设失败: {resp.text}"
    data = resp.json()["data"]
    hypothesis_id = data["id"]
    paths = data.get("paths", [])
    assert len(paths) >= 1, f"路径解析为空: {data}"
    print(f"[OK] 假设创建成功，ID: {hypothesis_id}，路径数: {len(paths)}")
    return hypothesis_id


def test_generate_data(project_id: str) -> None:
    """5. 数据生成。"""
    _print_section("5. 数据生成")
    resp = requests.post(
        f"{BASE_URL}/simulation/{project_id}/generate",
        json={"sample_size": 200},
        headers=HEADERS,
        timeout=60,
    )
    _print_response(resp)
    assert resp.status_code == 200, f"数据生成失败: {resp.text}"
    print("[OK] 数据生成完成")


def test_analyze_report(project_id: str) -> str:
    """6. 生成报告（调用 LLM R4 诊断）。"""
    _print_section("6. 生成报告")
    print("  调用 LLM R4，预计 10~30s...")
    t0 = time.time()
    resp = requests.post(
        f"{BASE_URL}/report/analyze/{project_id}",
        json={},
        headers=HEADERS,
        timeout=180,
    )
    print(f"  耗时: {time.time() - t0:.1f}s")
    _print_response(resp)
    assert resp.status_code == 200, f"报告生成失败: {resp.text}"
    data = resp.json()["data"]

    # 验证信效度结果
    reliability = data.get("reliability_results", [])
    assert len(reliability) >= 1, "信效度结果为空"
    print(f"[OK] 信效度结果: {len(reliability)} 维度")

    # 验证差异检验结果（P1-1 新功能）
    diff_tests = data.get("diff_tests")
    assert diff_tests is not None, "差异检验结果 diff_tests 缺失"
    assert isinstance(diff_tests, list), f"diff_tests 类型错误: {type(diff_tests)}"
    if diff_tests:
        first = diff_tests[0]
        for key in ("predictor", "outcome", "method", "statistic", "p_value"):
            assert key in first, f"diff_tests[0] 缺少字段 {key}: {first}"
        print(f"[OK] 差异检验: {len(diff_tests)} 条路径")
        for t in diff_tests:
            sig = "显著" if t.get("significant") else "不显著"
            p_val = t.get("p_value")
            p_str = f"{p_val:.4f}" if p_val is not None else "N/A"
            method = t.get("method_name") or t.get("method") or "—"
            print(f"    - {t.get('predictor')}→{t.get('outcome')} "
                  f"({method}) p={p_str} {sig}")
    else:
        print("⚠ 差异检验为空（可能项目未配置假设路径）")

    # 验证诊断结果
    diagnosis = data.get("diagnosis")
    assert diagnosis is not None, "诊断结果缺失"
    print(f"[OK] 诊断结果: passed={diagnosis.get('passed')}，"
          f"问题数={len(diagnosis.get('issues', []))}")

    report_id = data["id"]
    print(f"[OK] 报告生成完成，ID: {report_id}")
    return report_id


def test_get_report(project_id: str) -> None:
    """7. 获取报告（验证不落库字段 diff_tests 仍可获取）。"""
    _print_section("7. 获取报告（GET）")
    resp = requests.get(
        f"{BASE_URL}/report/{project_id}",
        headers=HEADERS,
        timeout=30,
    )
    _print_response(resp)
    assert resp.status_code == 200, f"获取报告失败: {resp.text}"
    data = resp.json()["data"]

    # 验证不落库字段 diff_tests 通过 GET 端点仍可获取（实时计算）
    diff_tests = data.get("diff_tests")
    assert diff_tests is not None, "GET 报告 diff_tests 缺失"
    print(f"[OK] GET 报告含 diff_tests: {len(diff_tests)} 条路径")

    # 验证信效度结果存在
    assert "reliability_results" in data
    assert "diagnosis" in data
    print(f"[OK] GET 报告字段完整")


def test_export_word(report_id: str) -> None:
    """8. 导出 Word（验证含差异检验章节，P1-3 新功能）。"""
    _print_section("8. 导出 Word")
    resp = requests.post(
        f"{BASE_URL}/report/export/{report_id}",
        json={"format": "word"},
        headers=HEADERS,
        timeout=60,
    )
    print(f"  状态码: {resp.status_code}")
    print(f"  Content-Type: {resp.headers.get('Content-Type')}")
    print(f"  Content-Disposition: {resp.headers.get('Content-Disposition')}")
    assert resp.status_code == 200, f"Word 导出失败: {resp.text}"
    assert "wordprocessingml" in resp.headers.get("Content-Type", "")
    assert len(resp.content) > 5000, f"Word 文件过小: {len(resp.content)} bytes"

    # 解析 docx 验证含差异检验章节
    try:
        from docx import Document
        doc = Document(io.BytesIO(resp.content))
        headings = [p.text for p in doc.paragraphs if p.style.name.startswith("Heading")]
        has_diff = any("差异分析" in h or "假设检验" in h for h in headings)
        assert has_diff, f"Word 缺少差异检验章节，现有章节: {headings}"
        print(f"[OK] Word 含差异检验章节: {len(resp.content)} bytes")
        print(f"  章节列表: {headings}")
    except ImportError:
        print("  [跳过 docx 解析] python-docx 未安装")


def test_export_excel(report_id: str) -> None:
    """9. 导出 Excel（验证含差异检验 sheet，P1-3 新功能）。"""
    _print_section("9. 导出 Excel")
    resp = requests.post(
        f"{BASE_URL}/report/export/{report_id}",
        json={"format": "excel"},
        headers=HEADERS,
        timeout=60,
    )
    print(f"  状态码: {resp.status_code}")
    print(f"  Content-Type: {resp.headers.get('Content-Type')}")
    assert resp.status_code == 200, f"Excel 导出失败: {resp.text}"
    assert "spreadsheetml" in resp.headers.get("Content-Type", "")
    assert len(resp.content) > 3000, f"Excel 文件过小: {len(resp.content)} bytes"

    # 解析 xlsx 验证含差异检验 sheet
    try:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(resp.content))
        sheets = wb.sheetnames
        assert "差异检验" in sheets, f"Excel 缺少差异检验 sheet，现有: {sheets}"
        ws = wb["差异检验"]
        headers = [c.value for c in ws[1]]
        print(f"[OK] Excel 含差异检验 sheet: {len(resp.content)} bytes")
        print(f"  sheets: {sheets}")
        print(f"  差异检验表头: {headers}")
        print(f"  差异检验数据行数: {ws.max_row - 1}")
    except ImportError:
        print("  [跳过 xlsx 解析] openpyxl 未安装")


def test_cleanup(project_id: str) -> None:
    """10. 清理测试项目。"""
    _print_section("10. 清理项目")
    resp = requests.delete(
        f"{BASE_URL}/projects/{project_id}",
        headers=HEADERS,
    )
    print(f"  状态码: {resp.status_code}")
    assert resp.status_code == 204, f"删除项目失败: {resp.text}"
    print(f"[OK] 项目已删除: {project_id}")


# ============================================================
# 异常分支测试（不依赖 LLM，可独立快速运行）
# ============================================================

def test_no_auth() -> None:
    """A. 未认证访问 → 401。"""
    _print_section("异常 A: 未认证访问")
    resp = requests.get(f"{BASE_URL}/projects/")
    _print_response(resp)
    assert resp.status_code == 401, f"期望 401，实际 {resp.status_code}"
    print("[OK] 未认证访问返回 401")


def test_invalid_token() -> None:
    """B. 错误 token → 401。"""
    _print_section("异常 B: 错误 token")
    resp = requests.get(
        f"{BASE_URL}/projects/",
        headers={"Authorization": "Bearer invalid-token"},
    )
    _print_response(resp)
    assert resp.status_code == 401, f"期望 401，实际 {resp.status_code}"
    print("[OK] 错误 token 返回 401")


def test_get_nonexistent_project() -> None:
    """C. 访问不存在的项目 → 404。"""
    _print_section("异常 C: 不存在的项目")
    fake_id = str(uuid.uuid4())
    resp = requests.get(
        f"{BASE_URL}/projects/{fake_id}",
        headers=HEADERS,
    )
    _print_response(resp)
    assert resp.status_code == 404, f"期望 404，实际 {resp.status_code}"
    print("[OK] 不存在的项目返回 404")


def test_status_guard_analyze_draft() -> None:
    """D. draft 项目直接 analyze → 400（需先激活套餐以绕过付费守卫）。"""
    _print_section("异常 D: draft 状态调用 analyze")
    _activate_dev_subscription("single")
    project_id = _create_draft_project("draft-analyze")
    try:
        resp = requests.post(
            f"{BASE_URL}/report/analyze/{project_id}",
            json={},
            headers=HEADERS,
            timeout=30,
        )
        _print_response(resp)
        assert resp.status_code == 400, f"期望 400，实际 {resp.status_code}"
        assert "状态" in resp.json().get("message", "") or "状态" in resp.text
        print("[OK] draft 状态调用 analyze 返回 400")
    finally:
        _delete_project(project_id)


def test_status_guard_export_data_draft() -> None:
    """E. draft 项目调用 export-data → 400（需先激活套餐以绕过付费守卫）。"""
    _print_section("异常 E: draft 状态调用 export-data")
    _activate_dev_subscription("single")
    project_id = _create_draft_project("draft-export")
    try:
        resp = requests.post(
            f"{BASE_URL}/simulation/{project_id}/export-data",
            json={"format": "excel"},
            headers=HEADERS,
            timeout=30,
        )
        _print_response(resp)
        assert resp.status_code == 400, f"期望 400，实际 {resp.status_code}"
        print("[OK] draft 状态调用 export-data 返回 400")
    finally:
        _delete_project(project_id)


def test_generate_nonexistent_project() -> None:
    """F. 不存在的 project_id 调用 generate → 404（需先激活套餐以绕过付费守卫）。"""
    _print_section("异常 F: 不存在的 project_id")
    _activate_dev_subscription("single")
    fake_id = str(uuid.uuid4())
    resp = requests.post(
        f"{BASE_URL}/simulation/{fake_id}/generate",
        json={"sample_size": 100},
        headers=HEADERS,
        timeout=30,
    )
    _print_response(resp)
    assert resp.status_code == 404, f"期望 404，实际 {resp.status_code}"
    print("[OK] 不存在的 project_id 返回 404")


def test_export_invalid_format() -> None:
    """G. 不存在的 report_id 导出 → 404（需先激活套餐以绕过付费守卫）。"""
    _print_section("异常 G: 不支持的导出格式")
    _activate_dev_subscription("single")
    fake_report_id = str(uuid.uuid4())
    resp = requests.post(
        f"{BASE_URL}/report/export/{fake_report_id}",
        json={"format": "pdf"},
        headers=HEADERS,
        timeout=30,
    )
    _print_response(resp)
    # 不存在的 report_id 先报 404
    assert resp.status_code == 404, f"期望 404，实际 {resp.status_code}"
    print("[OK] 不存在的 report_id 返回 404（format 校验未触发）")


def test_get_nonexistent_report() -> None:
    """H. 获取不存在的报告 → 404。"""
    _print_section("异常 H: 获取不存在的报告")
    project_id = _create_draft_project("no-report")
    try:
        resp = requests.get(
            f"{BASE_URL}/report/{project_id}",
            headers=HEADERS,
            timeout=30,
        )
        _print_response(resp)
        assert resp.status_code == 404, f"期望 404，实际 {resp.status_code}"
        print("[OK] 项目无报告返回 404")
    finally:
        _delete_project(project_id)


def test_paid_endpoint_requires_auth() -> None:
    """I. 未认证访问付费端点 → 401（认证先于权限校验）。"""
    _print_section("异常 I: 未认证访问付费端点")
    fake_id = str(uuid.uuid4())
    resp = requests.post(
        f"{BASE_URL}/simulation/{fake_id}/hypothesis",
        json={"raw_text": "测试"},
    )
    _print_response(resp)
    assert resp.status_code == 401, f"期望 401，实际 {resp.status_code}"
    print("[OK] 未认证访问付费端点返回 401")


# ============================================================
# 主入口
# ============================================================

def run_happy_path() -> tuple:
    """运行主流程（happy path）。

    Returns:
        (project_id, success): project_id 用于清理；success 表示主流程是否完整通过。
    """
    print("\n" + "#" * 60)
    print("# 主流程（happy path）")
    print("#" * 60)

    project_id = None
    try:
        test_health()
        project_id = test_create_project()
        test_inspect_questions(project_id)
        # 激活 single 套餐以通过后续付费端点（创建假设 / 生成 / 导出）
        _activate_dev_subscription("single")
        hypothesis_id = test_create_hypothesis(project_id)
        test_generate_data(project_id)
        report_id = test_analyze_report(project_id)
        test_get_report(project_id)
        test_export_word(report_id)
        test_export_excel(report_id)
        print("\n" + "=" * 60)
        print("[OK] 主流程全部通过")
        print("=" * 60)
        return project_id, True
    except AssertionError as e:
        print(f"\n[FAIL] 主流程失败: {e}")
        return project_id, False
    except Exception as e:
        print(f"\n[FAIL] 主流程异常: {type(e).__name__}: {e}")
        # 提示 LLM 连接问题
        if "Connection" in str(e) or "connection" in str(e):
            print("  提示：LLM 连接错误，请检查 .env 中 DEEPSEEK_API_KEY 是否为真实有效值")
        return project_id, False


def run_anomalies() -> int:
    """运行异常分支测试。返回失败数。"""
    print("\n" + "#" * 60)
    print("# 异常分支（anomalies）")
    print("#" * 60)

    tests = [
        test_no_auth,
        test_invalid_token,
        test_get_nonexistent_project,
        test_status_guard_analyze_draft,
        test_status_guard_export_data_draft,
        test_generate_nonexistent_project,
        test_export_invalid_format,
        test_get_nonexistent_report,
        test_paid_endpoint_requires_auth,
    ]

    failures = 0
    for test_fn in tests:
        try:
            test_fn()
        except AssertionError as e:
            print(f"[FAIL] {test_fn.__name__} 失败: {e}")
            failures += 1
        except Exception as e:
            print(f"[FAIL] {test_fn.__name__} 异常: {type(e).__name__}: {e}")
            failures += 1

    print("\n" + "=" * 60)
    if failures == 0:
        print("[OK] 异常分支全部通过")
    else:
        print(f"[FAIL] 异常分支失败: {failures}/{len(tests)}")
    print("=" * 60)
    return failures


def main() -> int:
    """运行全部测试。返回退出码（0=成功，1=失败）。"""
    print("\n" + "=" * 60)
    print("端到端测试：完整流程 + 异常分支")
    print("=" * 60)

    # 1. 主流程
    project_id, happy_success = run_happy_path()

    # 2. 异常分支（无论主流程成功与否都运行，因为异常分支不依赖 LLM）
    failures = run_anomalies()

    # 3. 清理主流程项目（即使失败也要清理）
    if project_id:
        _print_section("最终清理：删除主流程项目")
        _delete_project(project_id)
        print(f"[OK] 已清理主流程项目: {project_id}")

    # 4. 汇总
    print("\n" + "=" * 60)
    if happy_success and failures == 0:
        print("[OK] 全部测试通过")
        print("=" * 60)
        return 0
    elif not happy_success and failures == 0:
        print("⚠ 主流程未通过（可能是 LLM 连接问题），异常分支全部通过")
        print("=" * 60)
        return 1
    else:
        print(f"[FAIL] 测试失败（主流程: {'通过' if happy_success else '未通过'}，"
              f"异常分支失败数: {failures}）")
        print("=" * 60)
        return 1


if __name__ == "__main__":
    sys.exit(main())
