"""题目体检路由（题型识别 / 维度归属 / 反向题标记）。"""
import io
import os
from typing import Dict, List, Optional, Set
from uuid import UUID

from fastapi import APIRouter, Depends, File, Query, UploadFile, Request
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.database import get_db
from app.core.dependencies import get_current_user
from app.core.responses import ResponseModel
from app.core.exceptions import NotFoundException, ValidationException
from app.models.question import Question
from app.schemas.questionnaire import (
    QuestionInspectRequest,
    QuestionnaireStructure,
    QuestionResponse,
    QuestionUpdateRequest,
    QuestionUploadResponse,
    DimensionsResponse,
    DimensionUpdateRequest,
)
from app.services.inspector import inspect as inspect_service
from app.services.project_service import get_owned_project, update_project_status
from app.services.audit_service import AuditService, ACTION_TYPES
from app.services.questionnaire_health import QuestionnaireHealthEngine

router = APIRouter(prefix="/questionnaire", tags=["questionnaire"])


# 文件上传限制
_MAX_UPLOAD_SIZE = 2 * 1024 * 1024  # 2MB
_ALLOWED_EXTENSIONS = {".txt", ".docx", ".xlsx", ".pdf"}

# 扩展名 → 允许的 MIME 类型白名单（大小写不敏感）
_ALLOWED_MIME_TYPES: Dict[str, Set[str]] = {
    ".txt": {"text/plain", "application/octet-stream"},
    ".docx": {
        "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        "application/octet-stream",
    },
    ".xlsx": {
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/octet-stream",
    },
    ".pdf": {"application/pdf", "application/octet-stream"},
}

# 扩展名 → 文件头魔数（前 N 字节）；.txt 无固定魔数，跳过魔数校验
_FILE_MAGIC_BYTES: Dict[str, List[bytes]] = {
    # .docx / .xlsx 本质是 ZIP 容器，魔数为 PK\x03\x04
    ".docx": [b"PK\x03\x04"],
    ".xlsx": [b"PK\x03\x04"],
    ".pdf": [b"%PDF"],
}


def _validate_file_mime(filename: str, content_type: Optional[str]) -> None:
    """校验文件 MIME 类型与扩展名是否匹配。

    浏览器可能传 application/octet-stream（通用二进制），允许通过。
    """
    ext = os.path.splitext(filename)[1].lower()
    allowed = _ALLOWED_MIME_TYPES.get(ext)
    if not allowed:
        return  # 扩展名已在调用方校验，此处兜底放行

    if content_type and content_type not in allowed:
        raise ValidationException(
            f"文件 MIME 类型 {content_type} 与扩展名 {ext} 不匹配，疑似伪造文件"
        )


def _validate_file_magic(filename: str, content: bytes) -> None:
    """校验文件头魔数，防止扩展名伪装攻击。

    .txt 无固定魔数，跳过；其余格式按 _FILE_MAGIC_BYTES 校验。
    """
    ext = os.path.splitext(filename)[1].lower()
    magic_list = _FILE_MAGIC_BYTES.get(ext)
    if not magic_list:
        return  # .txt 等无魔数格式跳过

    # 文件内容不足以包含魔数 → 异常
    if len(content) < 4:
        raise ValidationException("文件内容过短，疑似损坏或伪造")

    for magic in magic_list:
        if content.startswith(magic):
            return

    raise ValidationException(
        f"文件头魔数与 {ext} 格式不匹配，疑似伪装文件"
    )


def _read_text_file(content: bytes) -> str:
    """读取文本文件，依次尝试 UTF-8 / GBK / latin1 编码。"""
    for encoding in ("utf-8", "gbk", "latin1"):
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise ValidationException("无法识别文本文件编码，请保存为 UTF-8 后重试")


def _read_docx_file(content: bytes) -> str:
    """读取 .docx 文件并提取段落文本。"""
    try:
        import docx
    except ImportError as exc:
        raise ValidationException("缺少 docx 解析依赖，请联系管理员安装 python-docx") from exc

    document = docx.Document(io.BytesIO(content))
    paragraphs = [p.text.strip() for p in document.paragraphs if p.text.strip()]
    return "\n".join(paragraphs)


def _read_excel_file(content: bytes) -> str:
    """读取 .xlsx 文件并提取首个工作表的文本内容。

    优先读取 A 列，如果 A 列为空则按行拼接所有非空单元格。
    """
    try:
        import openpyxl
    except ImportError as exc:
        raise ValidationException("缺少 Excel 解析依赖，请联系管理员安装 openpyxl") from exc

    workbook = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    sheet = workbook.active
    if not sheet:
        raise ValidationException("Excel 文件为空")

    lines = []
    first_col_values = []
    for row in sheet.iter_rows(values_only=True):
        first = row[0] if row else None
        if first is not None and str(first).strip():
            first_col_values.append(str(first).strip())

    if first_col_values:
        lines = first_col_values
    else:
        for row in sheet.iter_rows(values_only=True):
            row_text = " ".join(str(cell).strip() for cell in row if cell is not None and str(cell).strip())
            if row_text:
                lines.append(row_text)

    if not lines:
        raise ValidationException("Excel 文件未包含可识别的文本内容")

    return "\n".join(lines)


def _read_pdf_file(content: bytes) -> str:
    """读取 .pdf 文件并提取文本内容。"""
    try:
        from pypdf import PdfReader
    except ImportError as exc:
        raise ValidationException("缺少 PDF 解析依赖，请联系管理员安装 pypdf") from exc

    reader = PdfReader(io.BytesIO(content))
    if not reader.pages:
        raise ValidationException("PDF 文件为空")

    paragraphs = []
    for page in reader.pages:
        text = page.extract_text()
        if text:
            paragraphs.append(text.strip())

    full_text = "\n".join(paragraphs).strip()
    if not full_text:
        raise ValidationException("PDF 文件未包含可提取的文本（可能是扫描件或图片 PDF）")

    return full_text


@router.post(
    "/inspect",
    response_model=ResponseModel[QuestionnaireStructure],
    summary="题目体检",
    description="识别题型、维度归属、反向题，输出归属表。免费层能力，不含智能诊断。"
)
async def inspect(
    project_id: UUID,
    request: QuestionInspectRequest,
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """题目体检：识别题型、维度归属、反向题，输出归属表。

    免费层能力，不含智能诊断。
    """
    # 1. 验证项目存在且属于当前用户（含软删除过滤）
    project = await get_owned_project(db, project_id, current_user["id"])

    # 2. 调用体检服务（LLM 题型识别 / 维度归属 / 反向题）
    try:
        questionnaire_structure = inspect_service(request.text)
    except Exception as e:
        raise ValidationException(f"体检失败: {str(e)}")

    # 3. 保存题目到数据库
    for q in questionnaire_structure.questions:
        question = Question(
            project_id=project_id,
            index=q.index,
            text=q.text,
            question_type=q.question_type,
            dimension=q.dimension,
            is_reverse=q.is_reverse,
            confidence=q.confidence
        )
        db.add(question)

    # 4. 更新项目状态为 inspected
    update_project_status(project, "inspected", reason="题目体检完成")
    await db.flush()

    return ResponseModel(data=questionnaire_structure)


@router.post(
    "/upload",
    response_model=ResponseModel[QuestionUploadResponse],
    summary="上传问卷文件",
    description="上传 .txt / .docx / .xlsx / .pdf 文件并提取原始文本，单文件 ≤ 2MB。",
)
async def upload_questionnaire_file(
    project_id: UUID,
    file: UploadFile = File(..., description="问卷文件，支持 .txt / .docx / .xlsx / .pdf"),
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """上传问卷文件并提取原始文本。"""
    # 1. 验证项目存在且属于当前用户（含软删除过滤）
    await get_owned_project(db, project_id, current_user["id"])

    # 2. 校验文件名与扩展名
    if not file.filename:
        raise ValidationException("文件名不能为空")

    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in _ALLOWED_EXTENSIONS:
        raise ValidationException(f"不支持的文件格式：{ext}，仅支持 .txt / .docx / .xlsx / .pdf")

    # 3. 读取文件内容并校验大小
    content = await file.read()
    if len(content) > _MAX_UPLOAD_SIZE:
        raise ValidationException("文件大小超过 2MB 限制")

    if not content:
        raise ValidationException("文件内容为空")

    # 3.5 校验 MIME 类型与文件头魔数（防止伪装文件攻击）
    _validate_file_mime(file.filename, file.content_type)
    _validate_file_magic(file.filename, content)

    # 4. 提取文本
    try:
        if ext == ".txt":
            text = _read_text_file(content)
        elif ext == ".docx":
            text = _read_docx_file(content)
        elif ext == ".xlsx":
            text = _read_excel_file(content)
        elif ext == ".pdf":
            text = _read_pdf_file(content)
        else:
            raise ValidationException(f"不支持的文件格式：{ext}")
    except ValidationException:
        raise
    except Exception as e:
        raise ValidationException(f"文件解析失败: {str(e)}")

    return ResponseModel(data=QuestionUploadResponse(text=text))


@router.get(
    "/questions/{project_id}",
    response_model=ResponseModel[List[QuestionResponse]],
    summary="获取题目列表",
    description="获取项目的题目列表"
)
async def get_questions(
    project_id: UUID,
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """获取项目的题目列表。"""
    # 验证项目存在且属于当前用户（含软删除过滤）
    await get_owned_project(db, project_id, current_user["id"])

    # 查询题目
    result = await db.execute(
        select(Question)
        .where(Question.project_id == project_id)
        .order_by(Question.index)
    )
    questions = result.scalars().all()

    return ResponseModel(data=questions)


@router.patch(
    "/questions/{project_id}/{question_index}",
    response_model=ResponseModel[QuestionResponse],
    summary="更新单题（维度/反向题/置信度）",
    description="用户修正 AI 识别结果。仅允许更新 dimension/is_reverse/confidence。"
)
async def update_question(
    project_id: UUID,
    question_index: int,
    request: QuestionUpdateRequest,
    http_request: Request,
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """更新单题（用户修正 AI 识别结果）。"""
    # 1. 验证项目存在且属于当前用户（含软删除过滤）
    await get_owned_project(db, project_id, current_user["id"])

    # 2. 查询题目（按 project_id + index 定位）
    result = await db.execute(
        select(Question).where(
            Question.project_id == project_id,
            Question.index == question_index
        )
    )
    question = result.scalar_one_or_none()
    if not question:
        raise NotFoundException("题目不存在")

    # 3. 记录变更前的值（用于审计日志）
    old_dimension = question.dimension
    old_is_reverse = question.is_reverse

    # 4. 应用更新（仅非 None 字段）
    if request.dimension is not None:
        question.dimension = request.dimension
    if request.is_reverse is not None:
        question.is_reverse = request.is_reverse
    if request.confidence is not None:
        question.confidence = request.confidence

    # 5. 记录审计日志
    action_details = {"question_index": question_index}
    if request.dimension is not None and request.dimension != old_dimension:
        action_details["dimension_changed"] = {"from": old_dimension, "to": request.dimension}
        await AuditService.log_action(
            db=db,
            user_id=current_user["id"],
            action_type=ACTION_TYPES["DIMENSION_EDIT"],
            project_id=project_id,
            action_detail=action_details,
            ip_address=http_request.client.host if http_request.client else None,
            user_agent=http_request.headers.get("user-agent"),
        )
    if request.is_reverse is not None and request.is_reverse != old_is_reverse:
        action_details["reverse_changed"] = {"from": old_is_reverse, "to": request.is_reverse}
        await AuditService.log_action(
            db=db,
            user_id=current_user["id"],
            action_type=ACTION_TYPES["REVERSE_TOGGLE"],
            project_id=project_id,
            action_detail=action_details,
            ip_address=http_request.client.host if http_request.client else None,
            user_agent=http_request.headers.get("user-agent"),
        )

    await db.flush()
    return ResponseModel(data=question)


@router.get(
    "/dimensions/{project_id}",
    response_model=ResponseModel[DimensionsResponse],
    summary="获取维度列表",
    description="获取项目下所有题目的维度列表（去重，按出现顺序）。"
)
async def get_dimensions(
    project_id: UUID,
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """获取项目的维度列表。"""
    # 1. 验证项目存在且属于当前用户（含软删除过滤）
    await get_owned_project(db, project_id, current_user["id"])

    # 2. 查询维度并按出现顺序去重
    result = await db.execute(
        select(Question.dimension)
        .where(Question.project_id == project_id)
        .order_by(Question.index)
    )
    seen = set()
    dimensions = []
    for row in result.all():
        dim = row[0]
        if dim and dim not in seen:
            seen.add(dim)
            dimensions.append(dim)

    return ResponseModel(data=DimensionsResponse(dimensions=dimensions))


@router.post(
    "/dimensions/{project_id}",
    response_model=ResponseModel[DimensionsResponse],
    summary="新增/重命名维度",
    description="新增一个空维度，或重命名已有维度（同步更新所有相关题目）。"
)
async def update_dimensions(
    project_id: UUID,
    request: DimensionUpdateRequest,
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """新增或重命名维度。重命名会同步更新所有相关题目的 dimension 字段。"""
    # 1. 验证项目存在且属于当前用户（含软删除过滤）
    await get_owned_project(db, project_id, current_user["id"])

    # 2. 查询所有维度
    result = await db.execute(
        select(Question.dimension)
        .where(Question.project_id == project_id)
    )
    existing_dimensions = {row[0] for row in result.all() if row[0]}

    if request.action == "add":
        if request.name in existing_dimensions:
            raise ValidationException(f"维度 '{request.name}' 已存在")
        # 新增维度：暂时不绑定任何题目，仅返回更新后的列表
        existing_dimensions.add(request.name)

    elif request.action == "rename":
        if not request.old_name:
            raise ValidationException("重命名操作必须提供 old_name")
        if request.old_name not in existing_dimensions:
            raise ValidationException(f"原维度 '{request.old_name}' 不存在")
        if request.name in existing_dimensions and request.name != request.old_name:
            raise ValidationException(f"目标维度 '{request.name}' 已存在")

        # 同步更新所有相关题目
        result = await db.execute(
            select(Question).where(
                Question.project_id == project_id,
                Question.dimension == request.old_name
            )
        )
        questions_to_update = result.scalars().all()
        for q in questions_to_update:
            q.dimension = request.name

        existing_dimensions.discard(request.old_name)
        existing_dimensions.add(request.name)

    # 3. 返回按出现顺序排列的维度列表
    result = await db.execute(
        select(Question.dimension)
        .where(Question.project_id == project_id)
        .order_by(Question.index)
    )
    seen = set()
    dimensions = []
    for row in result.all():
        dim = row[0]
        if dim and dim not in seen:
            seen.add(dim)
            dimensions.append(dim)

    await db.flush()
    return ResponseModel(data=DimensionsResponse(dimensions=dimensions))


@router.post(
    "/wjx-import/{project_id}",
    response_model=ResponseModel[Dict],
    summary="导入问卷星导出文件",
    description="解析问卷星导出的 Excel/Word 文件，提取题目、选项、维度信息。",
)
async def import_wjx_file(
    project_id: UUID,
    file: UploadFile = File(..., description="问卷星导出的 Excel/Word 文件"),
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """导入问卷星导出文件。"""
    # 1. 验证项目存在且属于当前用户
    project = await get_owned_project(db, project_id, current_user["id"])

    # 2. 校验文件
    if not file.filename:
        raise ValidationException("文件名不能为空")

    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in {".xlsx", ".xls", ".docx"}:
        raise ValidationException(
            f"不支持的文件格式：{ext}，仅支持 .xlsx / .xls / .docx"
        )

    content = await file.read()
    if len(content) > _MAX_UPLOAD_SIZE:
        raise ValidationException("文件大小超过 2MB 限制")

    if not content:
        raise ValidationException("文件内容为空")

    # 3. 解析问卷星文件
    try:
        from app.services.wjx_parser import parse_wjx_export

        file_type = "excel" if ext in {".xlsx", ".xls"} else "word"
        parse_result = parse_wjx_export(content, file_type)

        questions = parse_result["questions"]
        dimensions = parse_result["dimensions"]
        warnings = parse_result["warnings"]

    except ValueError as e:
        raise ValidationException(str(e))
    except Exception as e:
        raise ValidationException(f"问卷星文件解析失败：{str(e)}")

    # 4. 保存题目到数据库
    for q in questions:
        question = Question(
            project_id=project_id,
            index=q["index"],
            text=q["text"],
            question_type=q["type"],
            dimension=q.get("dimension"),
            is_reverse=q.get("is_reverse", False),
            confidence=0.9,  # 问卷星导入的题目置信度较高
        )
        db.add(question)

    # 5. 更新项目状态为 inspected
    update_project_status(project, "inspected", reason="问卷星导入完成")
    await db.flush()

    return ResponseModel(
        data={
            "questions": questions,
            "dimensions": dimensions,
            "warnings": warnings,
            "question_count": len(questions),
            "dimension_count": len(dimensions),
        }
    )


# ---------------------------------------------------------------------------
# 问卷质量体检（纯规则引擎，不依赖 LLM）
# ---------------------------------------------------------------------------

@router.get("/{project_id}/health", response_model=ResponseModel)
async def get_questionnaire_health(
    project_id: UUID,
    db: AsyncSession = Depends(get_db),
    user=Depends(get_current_user),
):
    """对已识别的题目做质量诊断：题量/维度均衡/反向题/人口学/量表一致性/置信度/文本质量。

    纯规则引擎，不调用 LLM，响应快、零成本。需先完成题目识别（inspect）或问卷星导入。
    """
    project = await get_owned_project(db, project_id, user["id"])
    if not project:
        raise NotFoundException("项目不存在或无访问权限")

    result = await db.execute(
        select(Question)
        .where(Question.project_id == project_id)
        .order_by(Question.index)
    )
    questions = result.scalars().all()

    report = QuestionnaireHealthEngine(questions).run()
    return ResponseModel(data=report.to_dict())
